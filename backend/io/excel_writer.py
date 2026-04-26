from pathlib import Path
from typing import Dict, Optional, Tuple
from .helpers import add_serial_number
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _apply_bold_rows(ws, column_index: int = 1, keywords: list | None = None):
    """
    Bold entire rows where a specific column contains given keywords.
    """
    if keywords is None:
        keywords = ["Total"]

    for row in ws.iter_rows(min_row=2):  # skip header
        cell_value = row[column_index - 1].value

        if cell_value and any(k in str(cell_value) for k in keywords):
            for cell in row:
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=True,
                    color=cell.font.color,
                )


def _add_merged_title(ws, title: str, row: int = 1):
    """
    Adds a merged title row across the full width of the sheet.
    Should be called AFTER data is written (so max_column is correct).
    """

    max_col = ws.max_column
    last_col_letter = get_column_letter(max_col)
    ws.row_dimensions[row].height = 20

    # Merge cells across the row
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")

    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = Font(name="Times New Roman", bold=True, size=12, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Optional: add fill for professional look
    cell.fill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")


def _apply_styling(ws, header_rows=None):
    """
    Applies professional styling to a worksheet.
    If header_rows is provided, styles those specific rows as headers.
    """
    if header_rows is None:
        header_rows = [1]

    # Define styles
    header_font = Font("Times New Roman", bold=True, color="FFFFFF")
    content_font = Font(name="Times New Roman")
    header_fill = PatternFill(
        start_color="16365C", end_color="16365C", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Style header rows
    for row_idx in header_rows:
        for cell in ws[row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    header_rows_set = set(header_rows)
    # Style data cells and adjust column widths
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in col:
            if cell.row == 1:
                continue
            # Apply border to non-empty cells or cells within the table range
            if cell.value is not None:
                cell.border = thin_border
                if cell.row not in header_rows_set:
                    cell.font = content_font

            # Find max length for auto-width
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass

        # Adjust column width (with a bit of padding)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze the top row if it's a standard single-table sheet
    if len(header_rows) == 1 and header_rows[0] == 1:
        ws.freeze_panes = "A2"


TITLES = ["Reco Summary", ""]


def write_output_workbook(
    *,
    attachments: Dict[str, pd.DataFrame],
    rto_summary: pd.DataFrame,
    recon_summary: Tuple[pd.DataFrame, pd.DataFrame],
    dealership: str,
    output_dir: Optional[str] = None,
) -> str:
    """
    Writes final reconciliation output to an Excel workbook.
    Returns output file path.
    """

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"RTO_Reconciliation_{dealership}_{timestamp}.xlsx"

    output_path = Path(output_dir) / filename if output_dir else Path(filename)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Combined Reconciliation Summary
        summary_sheet_name = "Reco Summary"

        # 1. Delivery Summary
        recon_summary[0].to_excel(
            writer,
            sheet_name=summary_sheet_name,
            index=False,
            startrow=1,  # Leave row 0 for label
        )

        # 2. RTO Summary (Stacked)
        gap = 3
        second_header_row = 1 + len(recon_summary[0]) + 1 + gap
        recon_summary[1].to_excel(
            writer,
            sheet_name=summary_sheet_name,
            index=False,
            startrow=second_header_row + 1,  # Leave one row for label
        )

        # Add labels and style the combined sheet
        ws = writer.sheets[summary_sheet_name]
        _add_merged_title(ws, "Delivery Reco Summary", row=1)
        _add_merged_title(ws, "RTO Reco Summary", row=second_header_row + 1)

        # Attachments
        attachments["attachment_1"] = add_serial_number(attachments["attachment_1"])
        att_1_sheet_name = "Attachment 1"
        attachments["attachment_1"].to_excel(
            writer,
            sheet_name=att_1_sheet_name,
            index=False,
            startrow=2,
        )
        ws = writer.sheets[att_1_sheet_name]
        _add_merged_title(
            ws,
            "Attachment-1 : Deliveries Intimated but not found in RTO records",
        )

        attachments["attachment_2"] = add_serial_number(attachments["attachment_2"])
        att_2_sheet_name = "Attachement 2"
        attachments["attachment_2"].to_excel(
            writer, sheet_name=att_2_sheet_name, index=False, startrow=2
        )
        ws = writer.sheets[att_2_sheet_name]
        _add_merged_title(
            ws,
            "Attachment - 2: Found in RTO records but deliveries not Intimated",
        )

        if not attachments["delivery_reco"].empty and not attachments["rto_reco"].empty:
            del_reco_sheet_name = "Delivery Reco"
            attachments["delivery_reco"] = add_serial_number(
                attachments["delivery_reco"]
            )
            attachments["delivery_reco"].to_excel(
                writer,
                sheet_name="Delivery Reco",
                index=False,
                startrow=2,
            )
            ws = writer.sheets[del_reco_sheet_name]
            _add_merged_title(ws, "Delivery Reco: For Verifications of Results Only")
            rto_reco_sheet_name = "RTO Reco"
            attachments["rto_reco"] = add_serial_number(attachments["rto_reco"])
            attachments["rto_reco"].to_excel(
                writer,
                sheet_name="RTO Reco",
                index=False,
                startrow=2,
            )
            ws = writer.sheets[rto_reco_sheet_name]
            _add_merged_title(ws, "RTO Reco: For Verifications of Results Only")

        # RTO Summary (Standalone)
        rto_summary_sheet_name = "RTO Summary"
        rto_summary.to_excel(
            writer,
            sheet_name=rto_summary_sheet_name,
            index=False,
            startrow=2,
        )
        ws = writer.sheets[rto_summary_sheet_name]
        _add_merged_title(ws, "RTO Summary")

        # Apply styling to all sheets
        for name, sheet in writer.sheets.items():
            if name == summary_sheet_name:
                # Header rows: startrow + 1 (pandas writes header at startrow)
                # So row 2 (index 1) and row second_header_row + 2
                _apply_styling(sheet, header_rows=[2, second_header_row + 2])
                _apply_bold_rows(
                    sheet, column_index=1, keywords=["Total", "Matched", "Mismatch"]
                )  # Bold rows where first column has "Total"
            else:
                _apply_styling(sheet, header_rows=[3])
                _apply_bold_rows(sheet, column_index=1)

    return str(output_path)
